#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate personalized HRD follow-up feedback drafts from CSV data.

Core flow:
- read participant/action-plan rows from CSV
- generate a personalized follow-up draft for each participant with GLM or mock mode
- save progress so interrupted runs can resume safely
- write CSV manifest plus styled HTML email outputs

Recommended usage:
    python SCRIPT_glm_feedback_automation.py sample_data/교육_액션플랜_데이터.csv --mode mock
    python SCRIPT_glm_feedback_automation.py sample_data/교육_액션플랜_데이터.csv --mode glm
"""

from __future__ import annotations

import argparse
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import hashlib
import html
import json
import os
import re
import sys
import time
from pathlib import Path
from urllib import error, request


API_URL = "https://open.bigmodel.cn/api/coding/paas/v4/chat/completions"
DEFAULT_MODEL = "glm-4.5-air"
REQUEST_TIMEOUT = 180

DEFAULT_COURSE_NAME = "현대제철 리더십 교육 F/U"
DEFAULT_SENDER_NAME = "한충석 책임매니저"
DEFAULT_SENDER_ORG = "현대제철 컬처디자인팀"
DEFAULT_SUBJECT_PREFIX = "[현대제철 HRD]"

SYSTEM_PROMPT = """당신은 제조업 대기업 HRD 담당자를 돕는 교육 후속 커뮤니케이션 작성자다.
주어진 참가자 데이터를 바탕으로 개인화된 후속 메시지 초안을 작성하라.

반드시 지켜야 할 규칙:
1. 출력은 JSON 객체 하나만 반환한다.
2. JSON 키는 subject, message, next_action, tone_check를 사용한다.
3. message는 한국어 2~4문장, 300자 이내로 작성한다.
4. 과장된 칭찬, 근거 없는 성과 확신, 민감한 개인정보 언급을 피한다.
5. 액션플랜과 기대행동을 반영해 다음 행동이 분명히 보이게 쓴다.
6. tone_check에는 "ok" 또는 짧은 주의 메모만 넣는다.
7. HTML은 만들지 마라. HTML은 시스템이 별도로 렌더링한다."""

FIELD_ALIASES = {
    "participant_id": ["participant_id", "id", "참가자id", "응답자id", "사번"],
    "name": ["name", "이름", "성명"],
    "email": ["email", "이메일", "메일"],
    "company": ["company", "소속사", "회사"],
    "department": ["department", "부서", "소속"],
    "position": ["position", "직책", "직급", "직위"],
    "job_group": ["job_group", "직군", "직무군"],
    "aspiration": ["aspiration", "학습포부", "포부", "학습 포부"],
    "action_plan": ["action_plan", "액션플랜", "actionplan", "핵심실행계획", "핵심 실행계획"],
    "expected_behavior": ["expected_behavior", "기대행동", "기대 행동"],
    "training_comment": ["training_comment", "교육후코멘트", "교육 후 코멘트", "교육후 코멘트", "코멘트"],
    "top_takeaway": ["top_takeaway", "인상깊었던포인트", "인상 깊었던 포인트", "주요 인사이트"],
    "manager_name": ["manager_name", "관리자명", "매니저명", "강사명"],
}

REQUIRED_FIELDS = ("name", "action_plan")
OUTPUT_FIELDS = [
    "feedback_subject",
    "feedback_message",
    "feedback_next_action",
    "feedback_tone_check",
    "feedback_html",
    "feedback_html_path",
    "generation_status",
    "generation_engine",
    "generated_at",
]


def normalize_key(value: str) -> str:
    return re.sub(r"[\s_\-/()]+", "", value).lower()


def canonical_field_map(fieldnames: list[str]) -> dict[str, str]:
    normalized_to_actual = {normalize_key(name): name for name in fieldnames}
    mapping: dict[str, str] = {}

    for canonical_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            actual = normalized_to_actual.get(normalize_key(alias))
            if actual:
                mapping[canonical_name] = actual
                break

    return mapping


def read_input_rows(path: Path) -> tuple[list[dict[str, str]], list[str], dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV 헤더를 찾지 못했습니다.")
        raw_rows = [dict(row) for row in reader]
        fieldnames = list(reader.fieldnames)

    mapping = canonical_field_map(fieldnames)
    missing = [field for field in REQUIRED_FIELDS if field not in mapping]
    if missing:
        raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for row in raw_rows:
        normalized = {name: "" for name in FIELD_ALIASES}
        for canonical_name, actual_name in mapping.items():
            normalized[canonical_name] = (row.get(actual_name) or "").strip()
        normalized["_raw"] = row
        rows.append(normalized)

    return rows, fieldnames, mapping


def source_digest(rows: list[dict[str, str]], fieldnames: list[str]) -> str:
    payload = {
        "fieldnames": fieldnames,
        "rows": [{k: v for k, v in row.items() if k != "_raw"} for row in rows],
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def output_path_for(src_path: Path) -> Path:
    return src_path.with_name(f"{src_path.stem}_feedback_output.csv")


def progress_path_for(dst_path: Path) -> Path:
    return dst_path.with_name(f"{dst_path.name}.progress.json")


def preview_dir_for(dst_path: Path) -> Path:
    return dst_path.with_name(f"{dst_path.stem}_html_preview")


def index_html_path_for(dst_path: Path) -> Path:
    return dst_path.with_suffix(".html")


def xlsx_path_for(dst_path: Path) -> Path:
    return dst_path.with_suffix(".xlsx")


def load_progress(
    state_path: Path,
    src_path: Path,
    total_rows: int,
    digest: str,
) -> list[dict[str, str] | None]:
    if not state_path.exists():
        return []

    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"Ignoring unreadable progress file {state_path}: {exc}")
        return []

    if data.get("source") != str(src_path):
        print(f"Ignoring mismatched progress file {state_path}")
        return []
    if data.get("total_rows") != total_rows:
        print(f"Ignoring stale progress with different row count: {state_path}")
        return []
    if data.get("source_digest") != digest:
        print(f"Ignoring stale progress with different source digest: {state_path}")
        return []

    generated_rows = data.get("generated_rows", [])
    if not isinstance(generated_rows, list):
        print(f"Ignoring invalid progress payload: {state_path}")
        return []
    if len(generated_rows) > total_rows:
        print(f"Ignoring oversized progress payload: {state_path}")
        return []

    normalized: list[dict[str, str] | None] = []
    for item in generated_rows:
        if item is None:
            normalized.append(None)
        elif isinstance(item, dict):
            normalized.append({str(k): "" if v is None else str(v) for k, v in item.items()})
        else:
            normalized.append(None)

    normalized.extend([None] * (total_rows - len(normalized)))
    return normalized


def save_progress(
    state_path: Path,
    src_path: Path,
    total_rows: int,
    generated_rows: list[dict[str, str] | None],
    digest: str,
) -> None:
    payload = {
        "source": str(src_path),
        "source_digest": digest,
        "total_rows": total_rows,
        "completed_rows": sum(item is not None for item in generated_rows),
        "generated_rows": generated_rows,
    }
    state_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_messages(row: dict[str, str], args: argparse.Namespace) -> list[dict[str, str]]:
    participant_lines = [
        f"- 과정명: {args.course_name}",
        f"- 수신자 이름: {row.get('name') or '미상'}",
        f"- 소속사: {row.get('company') or '미입력'}",
        f"- 부서: {row.get('department') or '미입력'}",
        f"- 직책: {row.get('position') or '미입력'}",
        f"- 직군: {row.get('job_group') or '미입력'}",
        f"- 학습 포부: {row.get('aspiration') or '미입력'}",
        f"- 액션플랜: {row.get('action_plan') or '미입력'}",
        f"- 기대행동: {row.get('expected_behavior') or '미입력'}",
        f"- 교육 후 코멘트: {row.get('training_comment') or '미입력'}",
        f"- 교육 중 인상 깊었던 포인트: {row.get('top_takeaway') or '미입력'}",
        f"- 발신자(담당 매니저): {row.get('manager_name') or args.sender_name} / {args.sender_org}",
    ]
    prompt = "\n".join(participant_lines)
    return [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": prompt},
    ]


def call_glm(
    messages: list[dict[str, str]],
    model: str,
    temperature: float,
    api_key: str,
) -> str | None:
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
    }

    for attempt in range(3):
        try:
            req = request.Request(
                API_URL,
                data=json.dumps(payload).encode("utf-8"),
                headers=headers,
                method="POST",
            )
            with request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                body = resp.read().decode("utf-8")
                data = json.loads(body)
                return data["choices"][0]["message"]["content"]
        except error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="ignore")
            print(f"Error {exc.code}: {body}")
            if exc.code == 429:
                time.sleep(20 * (attempt + 1))
                continue
        except Exception as exc:
            print(f"Attempt {attempt + 1} failed: {exc}")
        time.sleep(5 * (attempt + 1))
    return None


def extract_json_object(text: str) -> dict[str, str] | None:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or start >= end:
        return None

    try:
        data = json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        return None

    if not isinstance(data, dict):
        return None

    return {str(k): "" if v is None else str(v).strip() for k, v in data.items()}


def sanitize_text(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", value or "").strip()
    return cleaned


def fallback_subject(row: dict[str, str], args: argparse.Namespace) -> str:
    return f"{args.subject_prefix} {row.get('name', '').strip()}님 교육 후속 실행 제안"


def build_mock_result(row: dict[str, str], args: argparse.Namespace) -> dict[str, str]:
    name = row.get("name") or "수강자"
    action_plan = sanitize_text(row.get("action_plan", "현업 적용 계획 정리"))
    expected_behavior = sanitize_text(row.get("expected_behavior", "팀 내 공유와 실행 점검"))
    comment = sanitize_text(row.get("training_comment", "실행 전 우선순위를 명확히 하고 싶다고 언급"))
    takeaway = sanitize_text(row.get("top_takeaway", "작게 시작해 빠르게 검증하는 방식"))
    comment_hint = re.split(r"(?<=[.!?])\s+", comment)[0].strip()
    message = (
        f"{name}님, 이번 주에는 '{action_plan}' 계획을 바로 실행할 수 있도록 첫 적용 범위를 작게 정해보시면 좋겠습니다. "
        f"교육에서 강조된 '{takeaway}' 포인트를 바탕으로 '{expected_behavior}'를 목표로 한 점검 기준 1개만 먼저 정해보시길 권합니다. "
        f"또한 '{comment_hint}' 의견을 반영해 메시지와 운영 방식을 더 간결하게 정리해보시기 바랍니다."
    )
    next_action = f"이번 주 내 첫 실행 항목 1개와 점검 기준 1개를 문서로 정리"
    return {
        "subject": fallback_subject(row, args),
        "message": sanitize_text(message),
        "next_action": next_action,
        "tone_check": "ok",
    }


def generate_feedback(row: dict[str, str], args: argparse.Namespace, api_key: str) -> dict[str, str] | None:
    if args.mode == "mock":
        return build_mock_result(row, args)

    response_text = call_glm(
        build_messages(row, args),
        model=args.model,
        temperature=args.temperature,
        api_key=api_key,
    )
    if response_text is None:
        return None

    payload = extract_json_object(response_text)
    if payload is None:
        return {
            "subject": fallback_subject(row, args),
            "message": sanitize_text(response_text)[:300],
            "next_action": sanitize_text(row.get("expected_behavior", "")),
            "tone_check": "JSON 파싱 실패",
        }

    return {
        "subject": sanitize_text(payload.get("subject") or fallback_subject(row, args)),
        "message": sanitize_text(payload.get("message") or ""),
        "next_action": sanitize_text(payload.get("next_action") or row.get("expected_behavior", "")),
        "tone_check": sanitize_text(payload.get("tone_check") or "ok"),
    }


def sanitize_filename(value: str) -> str:
    sanitized = re.sub(r"[^0-9A-Za-z가-힣._-]+", "_", value).strip("._")
    return sanitized or "preview"


def render_html(row: dict[str, str], result: dict[str, str], args: argparse.Namespace) -> str:
    """Render an Outlook-compatible HTML email using table-based layout + inline styles.

    Outlook uses Microsoft Word's rendering engine which does NOT support:
    - CSS variables (var())
    - CSS Grid / Flexbox
    - border-radius
    - backdrop-filter / glassmorphism
    - @keyframes animations
    - External web fonts via <link>
    - ::before / ::after pseudo-elements
    - <style> blocks are partially stripped

    This template uses ONLY:
    - Table-based layout
    - Inline styles on every element
    - MSO conditional comments for Outlook-specific behaviour
    - System-safe font stack
    - Simple solid backgrounds / borders
    """
    subject = html.escape(result["subject"])
    name = html.escape(row.get("name") or "수강자")
    company = html.escape(row.get("company") or "")
    department = html.escape(row.get("department") or "-")
    position = html.escape(row.get("position") or "-")
    action_plan = html.escape(row.get("action_plan") or "-")
    expected_behavior = html.escape(row.get("expected_behavior") or "-")
    takeaway = html.escape(row.get("top_takeaway") or "-")
    comment = html.escape(row.get("training_comment") or "-")
    body_sentences = [html.escape(sentence) for sentence in re.split(r"(?<=[.!?])\s+", result["message"]) if sentence]
    body_html = "\n".join(
        f'<p style="margin:0 0 14px 0;font-size:15px;line-height:1.9;color:#0e2233;">{s}</p>'
        for s in body_sentences
    )
    next_action = html.escape(result["next_action"] or "-")
    sender = html.escape(row.get("manager_name") or args.sender_name)
    sender_org = html.escape(args.sender_org)
    course_name = html.escape(args.course_name)
    company_dept = f"{company} {department}".strip()

    # Common inline font stack (Outlook-safe)
    FONT = "'Malgun Gothic', '맑은 고딕', 'Apple SD Gothic Neo', 'Segoe UI', Helvetica, Arial, sans-serif"

    return f"""<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml" lang="ko">
<head>
  <meta http-equiv="Content-Type" content="text/html; charset=utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{subject}</title>
  <!--[if mso]>
  <style type="text/css">
    table {{ border-collapse: collapse; }}
    td {{ font-family: {FONT}; }}
  </style>
  <![endif]-->
</head>
<body style="margin:0;padding:0;background-color:#eaf1f5;font-family:{FONT};font-size:15px;line-height:1.7;color:#0e2233;-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;">

  <!-- Outer wrapper table for centering -->
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#eaf1f5;">
    <tr>
      <td align="center" style="padding:40px 16px 50px 16px;">

        <!--[if mso]><table role="presentation" width="680" cellpadding="0" cellspacing="0" border="0" align="center"><tr><td><![endif]-->
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="max-width:680px;margin:0 auto;background-color:#ffffff;border:1px solid #d4e1e8;">

          <!-- ========== HEADER ========== -->
          <tr>
            <td style="padding:34px 36px 24px 36px;background-color:#f3f9f8;border-bottom:2px solid #0d8f84;">
              <!-- Eyebrow badge -->
              <table role="presentation" cellpadding="0" cellspacing="0" border="0">
                <tr>
                  <td style="background-color:#e6f4f2;padding:5px 14px;font-size:11px;font-weight:bold;color:#094d47;letter-spacing:0.08em;font-family:{FONT};">
                    &#9679;&nbsp; {course_name}
                  </td>
                </tr>
              </table>
              <!-- Title -->
              <h1 style="margin:16px 0 8px 0;font-size:24px;font-weight:bold;line-height:1.35;color:#091e2f;font-family:{FONT};">{subject}</h1>
              <!-- Recipient info -->
              <p style="margin:0;font-size:14px;line-height:1.7;color:#4a6a82;font-family:{FONT};">{name}&#xB2D8; &middot; {company_dept} &middot; {position}</p>
            </td>
          </tr>

          <!-- ========== BODY ========== -->
          <tr>
            <td style="padding:30px 36px 10px 36px;">
              <p style="margin:0 0 18px 0;font-size:15px;font-weight:bold;color:#094d47;font-family:{FONT};">{name}&#xB2D8; &#xC548;&#xB155;&#xD558;&#xC138;&#xC694;.</p>
              {body_html}
            </td>
          </tr>

          <!-- ========== INFO CARDS (2x2 table) ========== -->
          <tr>
            <td style="padding:0 36px 10px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
                <tr>
                  <!-- Card 1: Action Plan -->
                  <td width="50%" valign="top" style="padding:0 6px 12px 0;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xD575;&#xC2EC; &#xC2E4;&#xD589;&#xACC4;&#xD68D;</p>
                          <p style="margin:0;font-size:14px;line-height:1.8;color:#0e2233;font-family:{FONT};">{action_plan}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <!-- Card 2: Expected Behavior -->
                  <td width="50%" valign="top" style="padding:0 0 12px 6px;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xAE30;&#xB300; &#xD589;&#xB3D9;</p>
                          <p style="margin:0;font-size:14px;line-height:1.8;color:#0e2233;font-family:{FONT};">{expected_behavior}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
                <tr>
                  <!-- Card 3: Key Takeaway -->
                  <td width="50%" valign="top" style="padding:0 6px 12px 0;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xAD50;&#xC721; &#xD575;&#xC2EC; &#xD3EC;&#xC778;&#xD2B8;</p>
                          <p style="margin:0;font-size:14px;line-height:1.8;color:#0e2233;font-family:{FONT};">{takeaway}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <!-- Card 4: Comment -->
                  <td width="50%" valign="top" style="padding:0 0 12px 6px;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xCC38;&#xAC00;&#xC790; &#xCF54;&#xBA58;&#xD2B8;</p>
                          <p style="margin:0;font-size:14px;line-height:1.8;color:#0e2233;font-family:{FONT};">{comment}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- ========== NEXT ACTION ========== -->
          <tr>
            <td style="padding:6px 36px 20px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#0d8f84;">
                <tr>
                  <td style="padding:20px 24px;">
                    <p style="margin:0 0 6px 0;font-size:11px;font-weight:bold;letter-spacing:0.1em;color:#b8e6e0;font-family:{FONT};">NEXT ACTION</p>
                    <p style="margin:0;font-size:16px;line-height:1.8;font-weight:bold;color:#ffffff;font-family:{FONT};">{next_action}</p>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- ========== SIGNATURE ========== -->
          <tr>
            <td style="padding:0 36px 36px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-top:1px solid #dae4ec;">
                <tr>
                  <td style="padding:20px 0 0 0;font-size:13px;line-height:1.9;color:#4a6a82;font-family:{FONT};">
                    감사합니다.<br />
                    {sender}<br />
                    {sender_org}
                  </td>
                </tr>
              </table>
            </td>
          </tr>

        </table>
        <!--[if mso]></td></tr></table><![endif]-->

      </td>
    </tr>
  </table>

</body>
</html>
"""


def write_index_html(
    index_path: Path,
    rows: list[dict[str, str]],
    generated_rows: list[dict[str, str] | None],
    preview_dir: Path | None,
    args: argparse.Namespace,
) -> None:
    cards: list[str] = []

    for row, generated in zip(rows, generated_rows):
        if generated is None:
            cards.append(
                f"""
        <article class="result-card pending">
          <div class="card-top">
            <div>
              <p class="card-name">{html.escape(row.get("name") or "미상")}</p>
              <p class="card-meta">{html.escape(row.get("department") or "-")} | {html.escape(row.get("position") or "-")}</p>
            </div>
            <span class="status pending">Pending</span>
          </div>
        </article>
"""
            )
            continue

        relative_link = ""
        absolute_path = generated.get("feedback_html_path", "")
        if absolute_path and preview_dir is not None:
            relative_link = f"{preview_dir.name}/{Path(absolute_path).name}"

        cards.append(
            f"""
        <article class="result-card">
          <div class="card-top">
            <div>
              <p class="card-name">{html.escape(row.get("name") or "미상")}</p>
              <p class="card-meta">{html.escape(row.get("email") or "-")}</p>
              <p class="card-meta">{html.escape(row.get("department") or "-")} | {html.escape(row.get("position") or "-")}</p>
            </div>
            <span class="status completed">Ready</span>
          </div>
          <h2 class="card-subject">{html.escape(generated.get("feedback_subject", "-"))}</h2>
          <p class="card-body">{html.escape(generated.get("feedback_message", "-"))}</p>
          <div class="card-action">
            <span class="action-label">Next Action</span>
            <span class="action-value">{html.escape(generated.get("feedback_next_action", "-"))}</span>
          </div>
          {"<a class='card-link' href='" + html.escape(relative_link) + "' target='_blank'>개별 HTML 열기</a>" if relative_link else ""}
        </article>
"""
        )

    total_count = len(rows)
    completed_count = sum(item is not None for item in generated_rows)
    generated_at = datetime.now().isoformat(timespec="seconds")

    progress_pct = round(completed_count / total_count * 100) if total_count else 0

    index_html = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(args.course_name)} — 피드백 대시보드</title>
  <link rel="preconnect" href="https://cdn.jsdelivr.net">
  <link href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css" rel="stylesheet">
  <style>
    :root {{
      --bg: linear-gradient(168deg, #e8f0f4 0%, #f0f6f8 40%, #fafcfd 100%);
      --card: rgba(255, 255, 255, 0.78);
      --card-border: rgba(13, 143, 132, 0.08);
      --text: #0e2233;
      --muted: #5a7a92;
      --line: #dae4ec;
      --accent: #0d8f84;
      --accent-deep: #094d47;
      --accent-soft: rgba(13, 143, 132, 0.08);
      --accent-glow: rgba(13, 143, 132, 0.12);
      --shadow-sm: 0 2px 8px rgba(14, 34, 51, 0.04);
      --shadow-md: 0 12px 36px rgba(14, 34, 51, 0.07);
      --shadow-lg: 0 24px 64px rgba(14, 34, 51, 0.09);
      --radius-sm: 14px;
      --radius-md: 20px;
      --radius-lg: 28px;
    }}

    @keyframes fadeUp {{
      from {{ opacity: 0; transform: translateY(20px); }}
      to   {{ opacity: 1; transform: translateY(0); }}
    }}
    @keyframes slideIn {{
      from {{ opacity: 0; transform: translateY(12px); }}
      to   {{ opacity: 1; transform: translateY(0); }}
    }}
    @keyframes progressFill {{
      from {{ width: 0; }}
      to   {{ width: {progress_pct}%; }}
    }}

    * {{ box-sizing: border-box; margin: 0; padding: 0; }}

    body {{
      background: var(--bg);
      color: var(--text);
      font-family: "Pretendard Variable", "Pretendard", -apple-system, "Segoe UI", "Malgun Gothic", sans-serif;
      -webkit-font-smoothing: antialiased;
      -moz-osx-font-smoothing: grayscale;
    }}

    .page {{
      max-width: 1280px;
      margin: 0 auto;
      padding: 44px 24px 56px;
    }}

    /* -------- Hero -------- */
    .hero {{
      position: relative;
      padding: 36px 38px 32px;
      border-radius: var(--radius-lg);
      background: linear-gradient(135deg, #0d8f84 0%, #0a6b63 45%, #094d47 100%);
      color: #ffffff;
      box-shadow: var(--shadow-lg), inset 0 1px 0 rgba(255,255,255,0.08);
      overflow: hidden;
      animation: fadeUp 0.5s ease-out both;
    }}

    .hero::before {{
      content: "";
      position: absolute;
      top: -60px; right: -60px;
      width: 240px; height: 240px;
      border-radius: 50%;
      background: rgba(255, 255, 255, 0.04);
    }}
    .hero::after {{
      content: "";
      position: absolute;
      bottom: -80px; left: 30%;
      width: 300px; height: 300px;
      border-radius: 50%;
      background: rgba(255, 255, 255, 0.025);
    }}

    .hero-eyebrow {{
      position: relative;
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 5px 12px;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.1);
      border: 1px solid rgba(255, 255, 255, 0.12);
      font-size: 11px;
      font-weight: 700;
      letter-spacing: 0.1em;
      text-transform: uppercase;
    }}

    .hero h1 {{
      position: relative;
      margin: 14px 0 8px;
      font-size: 28px;
      font-weight: 800;
      line-height: 1.3;
      letter-spacing: -0.01em;
    }}

    .hero-desc {{
      position: relative;
      margin: 0;
      opacity: 0.82;
      line-height: 1.7;
      font-size: 14px;
    }}

    .stats {{
      position: relative;
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 14px;
      margin: 22px 0 0;
    }}

    .stat {{
      padding: 18px 20px;
      border-radius: var(--radius-md);
      background: rgba(255, 255, 255, 0.08);
      backdrop-filter: blur(6px);
      border: 1px solid rgba(255, 255, 255, 0.1);
      transition: background 0.25s ease;
    }}
    .stat:hover {{
      background: rgba(255, 255, 255, 0.14);
    }}

    .stat-label {{
      display: block;
      font-size: 11px;
      font-weight: 600;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      opacity: 0.68;
    }}

    .stat-value {{
      display: block;
      margin-top: 8px;
      font-size: 26px;
      font-weight: 800;
      letter-spacing: -0.02em;
    }}

    /* -------- Progress Bar -------- */
    .progress-bar {{
      position: relative;
      margin-top: 22px;
      height: 6px;
      border-radius: 3px;
      background: rgba(255, 255, 255, 0.12);
      overflow: hidden;
    }}
    .progress-fill {{
      height: 100%;
      border-radius: 3px;
      background: linear-gradient(90deg, rgba(255,255,255,0.4), rgba(255,255,255,0.7));
      animation: progressFill 1.2s ease-out 0.3s both;
    }}

    /* -------- Grid -------- */
    .grid {{
      display: grid;
      grid-template-columns: repeat(2, 1fr);
      gap: 18px;
      margin-top: 28px;
    }}

    .result-card {{
      padding: 24px;
      border-radius: var(--radius-md);
      background: var(--card);
      backdrop-filter: blur(16px) saturate(1.4);
      -webkit-backdrop-filter: blur(16px) saturate(1.4);
      border: 1px solid var(--card-border);
      box-shadow: var(--shadow-md);
      animation: slideIn 0.45s ease-out both;
      transition: transform 0.25s ease, box-shadow 0.25s ease;
    }}
    .result-card:hover {{
      transform: translateY(-4px);
      box-shadow: var(--shadow-lg);
    }}
    .result-card.pending {{
      opacity: 0.6;
    }}
    .result-card.pending:hover {{
      transform: none;
      box-shadow: var(--shadow-md);
    }}

    .card-top {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: flex-start;
    }}

    .card-name {{
      font-size: 19px;
      font-weight: 700;
      letter-spacing: -0.01em;
    }}

    .card-meta {{
      margin: 5px 0 0;
      font-size: 13px;
      line-height: 1.6;
      color: var(--muted);
    }}

    .status {{
      display: inline-flex;
      align-items: center;
      gap: 5px;
      padding: 5px 12px;
      border-radius: 999px;
      font-size: 11px;
      font-weight: 700;
      letter-spacing: 0.04em;
      text-transform: uppercase;
      white-space: nowrap;
    }}
    .status.completed {{
      background: var(--accent-soft);
      color: var(--accent-deep);
    }}
    .status.completed::before {{
      content: "\2713";
      font-size: 10px;
    }}
    .status.pending {{
      background: #edf1f5;
      color: #6b7f8e;
    }}
    .status.pending::before {{
      content: "\25CB";
      font-size: 10px;
    }}

    .card-subject {{
      margin: 16px 0 10px;
      font-size: 17px;
      font-weight: 700;
      line-height: 1.45;
      letter-spacing: -0.01em;
    }}

    .card-body {{
      font-size: 14px;
      line-height: 1.9;
      color: var(--text);
    }}

    .card-action {{
      display: grid;
      gap: 4px;
      margin-top: 18px;
      padding: 14px 18px;
      border-radius: var(--radius-sm);
      background: rgba(248, 251, 253, 0.7);
      border: 1px solid var(--line);
    }}

    .action-label {{
      font-size: 10px;
      font-weight: 700;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      color: var(--accent);
    }}

    .action-value {{
      font-size: 14px;
      line-height: 1.75;
      font-weight: 500;
    }}

    .card-link {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      margin-top: 18px;
      padding: 10px 18px;
      border-radius: 999px;
      background: linear-gradient(135deg, #0e2233 0%, #1a3a50 100%);
      color: #ffffff;
      text-decoration: none;
      font-size: 12.5px;
      font-weight: 700;
      letter-spacing: 0.02em;
      transition: background 0.2s ease, transform 0.15s ease;
    }}
    .card-link:hover {{
      background: linear-gradient(135deg, #1a3a50 0%, #0e2233 100%);
      transform: translateY(-1px);
    }}
    .card-link::after {{
      content: "\2192";
      font-size: 14px;
    }}

    /* -------- Footer -------- */
    .dashboard-footer {{
      margin-top: 36px;
      padding-top: 20px;
      border-top: 1px solid var(--line);
      text-align: center;
      font-size: 12px;
      color: var(--muted);
      line-height: 1.8;
    }}

    @media (max-width: 900px) {{
      .stats,
      .grid {{
        grid-template-columns: 1fr;
      }}
      .hero h1 {{
        font-size: 22px;
      }}
    }}

    @media print {{
      body {{ background: #ffffff; }}
      .hero {{ box-shadow: none; }}
      .result-card {{ box-shadow: none; border: 1px solid #ccc; page-break-inside: avoid; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <section class="hero">
      <span class="hero-eyebrow">{html.escape(args.course_name)}</span>
      <h1>개인화 후속 메일 대시보드</h1>
      <p class="hero-desc">생성 시각: {html.escape(generated_at)} &middot; 발신자: {html.escape(args.sender_name)} / {html.escape(args.sender_org)}</p>
      <div class="stats">
        <div class="stat">
          <span class="stat-label">전체 수신자</span>
          <span class="stat-value">{total_count}</span>
        </div>
        <div class="stat">
          <span class="stat-label">생성 완료</span>
          <span class="stat-value">{completed_count}</span>
        </div>
        <div class="stat">
          <span class="stat-label">Engine</span>
          <span class="stat-value">{html.escape(args.mode if args.mode == "mock" else args.model)}</span>
        </div>
      </div>
      <div class="progress-bar">
        <div class="progress-fill"></div>
      </div>
    </section>
    <section class="grid">
      {''.join(cards)}
    </section>
    <footer class="dashboard-footer">
      Generated by LETSAI HRD Feedback Automation &middot; {html.escape(generated_at)}
    </footer>
  </main>
</body>
</html>
"""
    index_path.write_text(index_html, encoding="utf-8")


def write_output_csv(
    dst_path: Path,
    original_fieldnames: list[str],
    rows: list[dict[str, str]],
    generated_rows: list[dict[str, str] | None],
    preview_dir: Path | None,
) -> None:
    fieldnames = original_fieldnames + [field for field in OUTPUT_FIELDS if field not in original_fieldnames]

    with dst_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row, generated in zip(rows, generated_rows):
            raw = dict(row["_raw"])
            if generated is None:
                for field in OUTPUT_FIELDS:
                    raw.setdefault(field, "")
            else:
                raw["feedback_subject"] = generated.get("feedback_subject", "")
                raw["feedback_message"] = generated.get("feedback_message", "")
                raw["feedback_next_action"] = generated.get("feedback_next_action", "")
                raw["feedback_tone_check"] = generated.get("feedback_tone_check", "")
                raw["feedback_html"] = generated.get("feedback_html", "")
                raw["feedback_html_path"] = generated.get("feedback_html_path", "")
                raw["generation_status"] = generated.get("generation_status", "")
                raw["generation_engine"] = generated.get("generation_engine", "")
                raw["generated_at"] = generated.get("generated_at", "")
            writer.writerow(raw)


def write_output_xlsx(
    xlsx_path: Path,
    original_fieldnames: list[str],
    rows: list[dict[str, str]],
    generated_rows: list[dict[str, str] | None],
) -> None:
    """Write an Excel workbook with the feedback results including HTML per row."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        print("openpyxl is not installed. Skipping xlsx output.")
        print("Install with: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Output"

    # Column definitions: (header label, field key, width)
    columns = [
        ("#", "_idx", 5),
        ("이름", "name", 12),
        ("소속사", "company", 18),
        ("부서", "department", 16),
        ("직책", "position", 12),
        ("이메일", "email", 28),
        ("메일 제목", "feedback_subject", 40),
        ("메일 본문", "feedback_message", 60),
        ("Next Action", "feedback_next_action", 40),
        ("톤 체크", "feedback_tone_check", 14),
        ("상태", "generation_status", 12),
        ("엔진", "generation_engine", 14),
        ("생성 시각", "generated_at", 20),
        ("HTML 전문", "feedback_html", 80),
    ]

    # Header styling
    header_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="0D8F84", end_color="0D8F84", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="DAE4EC"),
    )
    cell_font = Font(name="맑은 고딕", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)
    html_align = Alignment(vertical="top", wrap_text=False)

    # Write headers
    for col_idx, (label, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(columns)).column_letter}1"

    # Write data rows
    for row_idx, (row, generated) in enumerate(zip(rows, generated_rows), start=2):
        for col_idx, (_, field_key, _) in enumerate(columns, start=1):
            if field_key == "_idx":
                value = row_idx - 1
            elif generated is not None and field_key.startswith("feedback_") or field_key in ("generation_status", "generation_engine", "generated_at"):
                value = (generated or {}).get(field_key, "")
            else:
                value = row.get(field_key, "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border

            if field_key == "feedback_html":
                cell.alignment = html_align
            else:
                cell.alignment = cell_align

    # Alternate row coloring
    even_fill = PatternFill(start_color="F4FAFA", end_color="F4FAFA", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = even_fill

    try:
        wb.save(str(xlsx_path))
        print(f"Excel output: {xlsx_path}")
    except Exception as exc:
        print(f"Failed to save xlsx: {exc}")


def generate_one(
    index: int,
    row: dict[str, str],
    args: argparse.Namespace,
    api_key: str,
    preview_dir: Path | None,
) -> dict[str, str] | None:
    result = generate_feedback(row, args, api_key)
    if result is None:
        return None

    generated_at = datetime.now().isoformat(timespec="seconds")
    preview_path = ""
    html_content = render_html(row, result, args)

    if preview_dir is not None:
        preview_dir.mkdir(parents=True, exist_ok=True)
        preview_name = sanitize_filename(f"{index:03d}_{row.get('name', 'participant')}.html")
        html_path = preview_dir / preview_name
        html_path.write_text(html_content, encoding="utf-8")
        preview_path = str(html_path)

    return {
        "feedback_subject": result["subject"],
        "feedback_message": result["message"],
        "feedback_next_action": result["next_action"],
        "feedback_tone_check": result["tone_check"],
        "feedback_html": html_content,
        "feedback_html_path": preview_path,
        "generation_status": "completed",
        "generation_engine": args.mode if args.mode == "mock" else args.model,
        "generated_at": generated_at,
    }


def process_rows(
    src_path: Path,
    rows: list[dict[str, str]],
    original_fieldnames: list[str],
    args: argparse.Namespace,
    api_key: str,
) -> int:
    dst_path = args.output_csv or output_path_for(src_path)
    preview_dir = None if args.no_html_preview else (args.preview_dir or preview_dir_for(dst_path))
    index_html_path = None if args.no_html_preview else (args.output_html or index_html_path_for(dst_path))

    digest = source_digest(rows, original_fieldnames)
    state_path = progress_path_for(dst_path)
    generated_rows = load_progress(state_path, src_path, len(rows), digest)
    if not generated_rows:
        generated_rows = [None] * len(rows)

    completed = sum(item is not None for item in generated_rows)
    if completed:
        print(f"Resuming with {completed}/{len(rows)} completed rows.")

    pending_all = [(idx, row) for idx, row in enumerate(rows, start=1) if generated_rows[idx - 1] is None]
    pending = list(pending_all)
    if args.limit is not None:
        pending = pending[: args.limit]
    selected_indices = {idx for idx, _ in pending}

    if not pending:
        print("All rows already completed from progress file.")
        write_output_csv(dst_path, original_fieldnames, rows, generated_rows, preview_dir)
        if index_html_path is not None:
            write_index_html(index_html_path, rows, generated_rows, preview_dir, args)
        return 0

    print(f"Generating feedback for {len(pending)} rows...")

    if args.workers <= 1:
        for idx, row in pending:
            print(f"Processing row {idx}/{len(rows)}: {row.get('name', 'unknown')}")
            generated_rows[idx - 1] = generate_one(idx, row, args, api_key, preview_dir)
            save_progress(state_path, src_path, len(rows), generated_rows, digest)
            write_output_csv(dst_path, original_fieldnames, rows, generated_rows, preview_dir)
            time.sleep(0.2)
    else:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            future_to_index = {
                executor.submit(generate_one, idx, row, args, api_key, preview_dir): idx
                for idx, row in pending
            }
            for future in as_completed(future_to_index):
                idx = future_to_index[future]
                try:
                    generated_rows[idx - 1] = future.result()
                except Exception as exc:
                    print(f"Row {idx} raised {exc}")
                    generated_rows[idx - 1] = None
                completed_now = sum(item is not None for item in generated_rows)
                print(f"Completed row {idx}/{len(rows)} ({completed_now}/{len(rows)} done)")
                save_progress(state_path, src_path, len(rows), generated_rows, digest)
                write_output_csv(dst_path, original_fieldnames, rows, generated_rows, preview_dir)

    failed_selected = [idx for idx in selected_indices if generated_rows[idx - 1] is None]
    if failed_selected:
        print(f"Incomplete generation. Failed rows: {failed_selected}")
        save_progress(state_path, src_path, len(rows), generated_rows, digest)
        return 1

    remaining = [idx for idx, item in enumerate(generated_rows, start=1) if item is None]
    if args.limit is not None and len(pending) < len(pending_all):
        print(f"Partial run completed. Remaining rows: {remaining}")
        save_progress(state_path, src_path, len(rows), generated_rows, digest)
        write_output_csv(dst_path, original_fieldnames, rows, generated_rows, preview_dir)
        if index_html_path is not None:
            write_index_html(index_html_path, rows, generated_rows, preview_dir, args)
        return 0

    if state_path.exists():
        state_path.unlink()
    write_output_csv(dst_path, original_fieldnames, rows, generated_rows, preview_dir)
    if index_html_path is not None:
        write_index_html(index_html_path, rows, generated_rows, preview_dir, args)
    xlsx_out = xlsx_path_for(dst_path)
    write_output_xlsx(xlsx_out, original_fieldnames, rows, generated_rows)
    print(f"Finished! Saved to {dst_path}")
    if preview_dir is not None:
        print(f"HTML outputs: {preview_dir}")
    if index_html_path is not None:
        print(f"HTML index: {index_html_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate personalized HRD follow-up feedback drafts from CSV.")
    parser.add_argument("input_csv", type=Path, help="Input CSV path")
    parser.add_argument("--output-csv", type=Path, default=None, help="Output CSV path")
    parser.add_argument("--preview-dir", type=Path, default=None, help="Directory for HTML previews")
    parser.add_argument("--output-html", type=Path, default=None, help="Summary HTML output path")
    parser.add_argument(
        "--no-html-preview",
        "--no-html-output",
        dest="no_html_preview",
        action="store_true",
        help="Do not write styled HTML files",
    )
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output CSV")
    parser.add_argument("--mode", choices=["glm", "mock"], default="mock", help="Generation mode")
    parser.add_argument("--model", type=str, default=DEFAULT_MODEL, help="GLM model name")
    parser.add_argument("--temperature", type=float, default=0.3, help="GLM temperature")
    parser.add_argument("--workers", type=int, default=1, help="Concurrent workers")
    parser.add_argument("--limit", type=int, default=None, help="Process only the first N pending rows")
    parser.add_argument("--course-name", type=str, default=DEFAULT_COURSE_NAME, help="Course name")
    parser.add_argument("--sender-name", type=str, default=DEFAULT_SENDER_NAME, help="Sender name")
    parser.add_argument("--sender-org", type=str, default=DEFAULT_SENDER_ORG, help="Sender organization")
    parser.add_argument("--subject-prefix", type=str, default=DEFAULT_SUBJECT_PREFIX, help="Subject prefix")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if not args.input_csv.exists():
        print(f"Input CSV not found: {args.input_csv}")
        return 1

    api_key = os.getenv("API_KEY_ZAI", "").strip()
    if args.mode == "glm" and not api_key:
        print("API_KEY_ZAI is missing. Set API_KEY_ZAI or use --mode mock.")
        return 1

    output_csv = args.output_csv or output_path_for(args.input_csv)
    progress_path = progress_path_for(output_csv)
    if output_csv.exists() and not args.overwrite and not progress_path.exists():
        print(f"Output already exists: {output_csv}")
        print("Use --overwrite to regenerate.")
        return 1

    try:
        rows, original_fieldnames, _ = read_input_rows(args.input_csv)
    except Exception as exc:
        print(f"Failed to read input CSV: {exc}")
        return 1

    if not rows:
        print("No input rows found.")
        return 0

    print(f"Input rows: {len(rows)}")
    print(f"Mode: {args.mode}")
    print(f"Workers: {args.workers}")
    if args.limit:
        print(f"Limit: {args.limit}")

    return process_rows(args.input_csv, rows, original_fieldnames, args, api_key)


if __name__ == "__main__":
    raise SystemExit(main())
