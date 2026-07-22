#!/usr/bin/env python3
# 바로 실행 — 프로젝트 루트에서 아래 줄을 그대로 복사해 실행하십시오.
# python3 SCRIPT_glm_needs_email_automation.py "sample_data/HR_AI_교육참가자_니즈분석_20260722.csv" --mode mock
# python3 SCRIPT_glm_needs_email_automation.py "sample_data/HR_AI_교육참가자_니즈분석_20260722.csv" --mode glm
# -*- coding: utf-8 -*-
"""Generate personalized HRD AI seminar pre-invitation & needs-based HTML emails from CSV data.

Core flow:
- Read participant needs analysis rows from CSV (핵심니즈_추론, 심리상태_추론, 교육을통해얻고싶은것_추정)
- Generate personalized email text based on participant psychology & core needs via GLM or mock mode
- Save progress so interrupted runs can resume safely
- Write CSV manifest plus Outlook-compatible styled HTML email outputs and interactive index preview
"""

from __future__ import annotations

import argparse
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import hashlib
import html
import json
import os
import re
import sys
import time
from pathlib import Path
from urllib import error, request


def load_local_env(path: Path) -> None:
    """Load a project-local .env without overriding OS environment variables."""
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env(Path(__file__).with_name(".env"))

API_BASE_URL = os.getenv("GLM_BASE_URL", "https://api.z.ai/api/coding/paas/v4").rstrip("/")
API_URL = f"{API_BASE_URL}/chat/completions"
DEFAULT_MODEL = os.getenv("GLM_MODEL", "glm-4.5-air")
REQUEST_TIMEOUT = 180

DEFAULT_COURSE_NAME = "2026 HRD 부문의 AI 활용 세미나"
DEFAULT_SENDER_NAME = "한국인사관리협회 교육기획팀"
DEFAULT_SENDER_ORG = "HRD AI 세미나 운영본부"
DEFAULT_SUBJECT_PREFIX = "[HRD AI 세미나]"

SYSTEM_PROMPT = """당신은 HRD/HR 부문 AI 활용 세미나를 총괄하는 전문 교육 기획자다.
참가자의 사전 설문 및 추론 데이터(핵심 니즈, 심리 상태, 교육을 통해 얻고 싶은 점, AI 활용수준, 가장 큰 장애물 등)를 바탕으로,
참가자 개개인의 고민과 심리 상태에 깊이 공감하고, 세미나에서 해결할 핵심 가치와 기대 효과를 안내하며, 따뜻한 격려의 말을 전하는 맞춤형 사전 안내 이메일 문구를 작성하라.

반드시 지켜야 할 규칙:
1. 출력은 JSON 객체 하나만 반환한다.
2. JSON 키는 subject, message, encouragement, next_action, tone_check를 사용한다.
3. subject는 수신자의 소속/성명과 핵심 해결과제를 명시한 신뢰감 있는 제목으로 작성한다.
4. message는 3~5문장 (350자 이내)으로:
   - 참가자의 심리상태(예: 보안 우려, 기초역량 부족, ROI 증명 부담, 조직 설득 민감성 등)에 깊이 공감하고,
   - 핵심 니즈 및 교육을 통해 얻고자 하는 바(사례 BP, 보안 대안, 에이전트 제작, ROI 측정 등)가 세미나에서 어떻게 다뤄지는지 정중하고 명확하게 안내한다.
5. encouragement: 참가자의 적극적인 도전 의지와 역량 개발 노력을 진심으로 응원하고 사기를 북돋아주는 따뜻한 격려의 1~2문장(60~100자)을 작성한다.
6. next_action: 세미나 참석 전 사전 준비/지참 권장 사항 1~2개를 작성한다.
7. tone_check에는 "ok" 또는 짧은 검토 메모만 넣는다.
8. HTML 태그는 작성하지 마라. HTML은 시스템이 별도로 렌더링한다."""

FIELD_ALIASES = {
    "participant_id": ["번호", "participant_id", "id", "응답자id"],
    "name": ["성명", "이름", "name"],
    "email": ["수료증이메일", "이메일", "email", "메일"],
    "company": ["회사명", "소속사", "company", "회사"],
    "department": ["부서", "department", "소속"],
    "position": ["직급", "직책", "position"],
    "experience": ["HR업무경력", "경력", "experience"],
    "job_group": ["담당업무", "직무", "직군", "job_group"],
    "ai_level": ["Q1_AI활용수준", "AI활용수준", "활용수준"],
    "obstacle": ["Q2_가장큰장애물", "가장큰장애물", "장애물"],
    "learning_topic": ["Q3_학습희망주제", "학습희망주제", "희망주제"],
    "question": ["Q4_강사질문", "강사질문", "질문"],
    "industry": ["업종", "주요사업", "industry"],
    "core_needs": ["핵심니즈_추론", "핵심니즈", "핵심 니즈"],
    "psychology_state": ["심리상태_추론", "심리상태", "심리 상태"],
    "desired_outcome": ["교육을통해얻고싶은것_추정", "교육을통해얻고싶은것", "교육을 통해 얻고 싶은 것", "얻고싶은것"],
    "reasoning": ["추론근거"],
    "confidence": ["추론확신도"],
}

REQUIRED_FIELDS = ("name", "core_needs")
OUTPUT_FIELDS = [
    "feedback_subject",
    "feedback_message",
    "feedback_encouragement",
    "feedback_next_action",
    "feedback_tone_check",
    "feedback_html",
    "feedback_html_path",
    "generation_status",
    "generation_engine",
    "generated_at",
]


def normalize_key(value: str) -> str:
    return re.sub(r"[\s_\-/()]+", "", value).lower()


def canonical_field_map(fieldnames: list[str]) -> dict[str, str]:
    normalized_to_actual = {normalize_key(name): name for name in fieldnames}
    mapping: dict[str, str] = {}

    for canonical_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            actual = normalized_to_actual.get(normalize_key(alias))
            if actual:
                mapping[canonical_name] = actual
                break

    return mapping


def read_input_rows(path: Path) -> tuple[list[dict[str, str]], list[str], dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV 헤더를 찾지 못했습니다.")
        raw_rows = [dict(row) for row in reader]
        fieldnames = list(reader.fieldnames)

    mapping = canonical_field_map(fieldnames)
    missing = [field for field in REQUIRED_FIELDS if field not in mapping]
    if missing:
        raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for row in raw_rows:
        normalized = {name: "" for name in FIELD_ALIASES}
        for canonical_name, actual_name in mapping.items():
            normalized[canonical_name] = (row.get(actual_name) or "").strip()
        normalized["_raw"] = row
        rows.append(normalized)

    return rows, fieldnames, mapping


def source_digest(rows: list[dict[str, str]], fieldnames: list[str]) -> str:
    payload = {
        "fieldnames": fieldnames,
        "rows": [{k: v for k, v in row.items() if k != "_raw"} for row in rows],
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def output_path_for(src_path: Path) -> Path:
    return src_path.with_name(f"{src_path.stem}_feedback_output.csv")


def progress_path_for(dst_path: Path) -> Path:
    return dst_path.with_name(f"{dst_path.name}.progress.json")


def preview_dir_for(dst_path: Path) -> Path:
    return dst_path.with_name(f"{dst_path.stem}_html_preview")


def index_html_path_for(dst_path: Path) -> Path:
    return dst_path.with_suffix(".html")


def xlsx_path_for(dst_path: Path) -> Path:
    return dst_path.with_suffix(".xlsx")


def load_progress(
    state_path: Path,
    src_path: Path,
    total_rows: int,
    digest: str,
) -> list[dict[str, str] | None]:
    if not state_path.exists():
        return []

    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"Ignoring unreadable progress file {state_path}: {exc}")
        return []

    if data.get("source") != str(src_path):
        print(f"Ignoring mismatched progress file {state_path}")
        return []
    if data.get("total_rows") != total_rows:
        print(f"Ignoring stale progress with different row count: {state_path}")
        return []
    if data.get("source_digest") != digest:
        print(f"Ignoring stale progress with different source digest: {state_path}")
        return []

    generated_rows = data.get("generated_rows", [])
    if not isinstance(generated_rows, list):
        print(f"Ignoring invalid progress payload: {state_path}")
        return []
    if len(generated_rows) > total_rows:
        print(f"Ignoring oversized progress payload: {state_path}")
        return []

    normalized: list[dict[str, str] | None] = []
    for item in generated_rows:
        if item is None:
            normalized.append(None)
        elif isinstance(item, dict):
            normalized.append({str(k): "" if v is None else str(v) for k, v in item.items()})
        else:
            normalized.append(None)

    normalized.extend([None] * (total_rows - len(normalized)))
    return normalized


def save_progress(
    state_path: Path,
    src_path: Path,
    total_rows: int,
    generated_rows: list[dict[str, str] | None],
    digest: str,
) -> None:
    payload = {
        "source": str(src_path),
        "source_digest": digest,
        "total_rows": total_rows,
        "completed_rows": sum(item is not None for item in generated_rows),
        "generated_rows": generated_rows,
    }
    state_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_messages(row: dict[str, str], args: argparse.Namespace) -> list[dict[str, str]]:
    participant_lines = [
        f"- 과정명: {args.course_name}",
        f"- 수신자 이름: {row.get('name') or '미상'}",
        f"- 소속사: {row.get('company') or '미입력'}",
        f"- 부서/직급: {row.get('department') or '미입력'} / {row.get('position') or '미입력'}",
        f"- HR 경력/담당업무: {row.get('experience') or '-'} / {row.get('job_group') or '미입력'}",
        f"- AI 활용수준: {row.get('ai_level') or '미입력'}",
        f"- 주요 장애물: {row.get('obstacle') or '미입력'}",
        f"- 희망 학습 주제: {row.get('learning_topic') or '미입력'}",
        f"- 강사 사전 질문: {row.get('question') or '없음'}",
        f"- [핵심 분석] 핵심 니즈: {row.get('core_needs') or '미입력'}",
        f"- [핵심 분석] 심리 상태: {row.get('psychology_state') or '미입력'}",
        f"- [핵심 분석] 교육을 통해 얻고 싶은 것: {row.get('desired_outcome') or '미입력'}",
        f"- 발신자: {args.sender_name} / {args.sender_org}",
    ]
    prompt = "\n".join(participant_lines)
    return [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": prompt},
    ]


def call_glm(
    messages: list[dict[str, str]],
    model: str,
    temperature: float,
    api_key: str,
) -> str | None:
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
    }

    for attempt in range(3):
        try:
            req = request.Request(
                API_URL,
                data=json.dumps(payload).encode("utf-8"),
                headers=headers,
                method="POST",
            )
            with request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                body = resp.read().decode("utf-8")
                data = json.loads(body)
                return data["choices"][0]["message"]["content"]
        except error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="ignore")
            print(f"Error {exc.code}: {body}")
            if exc.code == 429:
                time.sleep(20 * (attempt + 1))
                continue
        except Exception as exc:
            print(f"Attempt {attempt + 1} failed: {exc}")
        time.sleep(5 * (attempt + 1))
    return None


def extract_json_object(text: str) -> dict[str, str] | None:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or start >= end:
        return None

    try:
        data = json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        return None

    if not isinstance(data, dict):
        return None

    return {str(k): "" if v is None else str(v).strip() for k, v in data.items()}


def sanitize_text(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", value or "").strip()
    return cleaned


def fallback_subject(row: dict[str, str], args: argparse.Namespace) -> str:
    company = row.get("company", "").strip()
    name = row.get("name", "").strip()
    position = row.get("position", "").strip()
    target = f"{company} {name} {position}".strip()
    return f"{args.subject_prefix} {target}님을 위한 사전 분석 및 세미나 가이드"


def build_mock_result(row: dict[str, str], args: argparse.Namespace) -> dict[str, str]:
    name = row.get("name") or "담당자"
    company = row.get("company") or "소속사"
    position = row.get("position") or "선생"
    core_needs = sanitize_text(row.get("core_needs", "HRD 업무에 맞춤화된 AI 활용 방안 확보"))
    psychology_state = sanitize_text(row.get("psychology_state", "도입 의지는 높으나 구체적 실행에 신중한 상태"))
    desired_outcome = sanitize_text(row.get("desired_outcome", "실무에 즉시 적용 가능한 템플릿과 성공 사례"))

    message = (
        f"{name} {position}님, 사전 설문을 통해 '{core_needs}'라는 핵심 니즈를 확인했습니다. "
        f"현재 '{psychology_state}'에 깊이 공감하며, 이번 세미나에서는 이러한 우려와 부담을 명확히 해소할 수 있는 실질적 솔루션을 제공하고자 합니다. "
        f"특히 요청해주신 '{desired_outcome}'을(를) 세미나 현장에서 직접 검증하고 바로 활용하실 수 있도록 실습형 커리큘럼으로 준비하였습니다."
    )
    encouragement = (
        f"{name} {position}님의 적극적인 학습 열정과 실무 혁신 의지가 {company}의 성공적인 AX 전환을 이끄는 가장 든든한 동력이 될 것입니다. "
        f"이번 세미나가 그 확실한 자신감과 계기를 드리는 시간이 되도록 최선을 다해 지원하겠습니다!"
    )
    next_action = "세미나 참석 전 현재 사용 중인 HRD 템플릿 1개 지참 및 사전 질문 준비"
    return {
        "subject": fallback_subject(row, args),
        "message": sanitize_text(message),
        "encouragement": sanitize_text(encouragement),
        "next_action": next_action,
        "tone_check": "ok",
    }


def generate_feedback(row: dict[str, str], args: argparse.Namespace, api_key: str) -> dict[str, str] | None:
    if args.mode == "mock":
        return build_mock_result(row, args)

    response_text = call_glm(
        build_messages(row, args),
        model=args.model,
        temperature=args.temperature,
        api_key=api_key,
    )
    if response_text is None:
        return None

    payload = extract_json_object(response_text)
    name = row.get("name") or "담당자"
    company = row.get("company") or "소속사"
    position = row.get("position") or ""
    default_encouragement = f"{name} {position}님의 끊임없는 역량 개발 노력이 {company}의 성공적인 AI 전환을 이끄는 큰 동력이 될 것입니다. 이번 세미나에서 힘을 실어드리겠습니다!"

    if payload is None:
        return {
            "subject": fallback_subject(row, args),
            "message": sanitize_text(response_text)[:350],
            "encouragement": sanitize_text(default_encouragement),
            "next_action": "사전 질문 및 소속사 보안 기준 사전 확인",
            "tone_check": "JSON 파싱 실패",
        }

    return {
        "subject": sanitize_text(payload.get("subject") or fallback_subject(row, args)),
        "message": sanitize_text(payload.get("message") or ""),
        "encouragement": sanitize_text(payload.get("encouragement") or default_encouragement),
        "next_action": sanitize_text(payload.get("next_action") or "사전 질문 및 소속사 보안 기준 사전 확인"),
        "tone_check": sanitize_text(payload.get("tone_check") or "ok"),
    }


def sanitize_filename(value: str) -> str:
    sanitized = re.sub(r"[^0-9A-Za-z가-힣._-]+", "_", value).strip("._")
    return sanitized or "preview"


def render_html(row: dict[str, str], result: dict[str, str], args: argparse.Namespace) -> str:
    """Render an Outlook-compatible HTML email with table-based layout and inline styles."""
    subject = html.escape(result["subject"])
    name = html.escape(row.get("name") or "담당자")
    company = html.escape(row.get("company") or "")
    department = html.escape(row.get("department") or "-")
    position = html.escape(row.get("position") or "-")

    core_needs = html.escape(row.get("core_needs") or "-")
    psychology_state = html.escape(row.get("psychology_state") or "-")
    desired_outcome = html.escape(row.get("desired_outcome") or "-")
    ai_level = html.escape(row.get("ai_level") or "-")
    obstacle = html.escape(row.get("obstacle") or "-")
    encouragement = html.escape(result.get("encouragement") or "-")

    body_sentences = [html.escape(sentence) for sentence in re.split(r"(?<=[.!?])\s+", result["message"]) if sentence]
    body_html = "\n".join(
        f'<p style="margin:0 0 14px 0;font-size:15px;line-height:1.9;color:#0e2233;">{s}</p>'
        for s in body_sentences
    )
    next_action = html.escape(result["next_action"] or "-")
    sender = html.escape(args.sender_name)
    sender_org = html.escape(args.sender_org)
    course_name = html.escape(args.course_name)
    company_dept = f"{company} {department}".strip()

    FONT = "'Malgun Gothic', '맑은 고딕', 'Apple SD Gothic Neo', 'Segoe UI', Helvetica, Arial, sans-serif"

    return f"""<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml" lang="ko">
<head>
  <meta http-equiv="Content-Type" content="text/html; charset=utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{subject}</title>
  <!--[if mso]>
  <style type="text/css">
    table {{ border-collapse: collapse; }}
    td {{ font-family: {FONT}; }}
  </style>
  <![endif]-->
</head>
<body style="margin:0;padding:0;background-color:#eaf1f5;font-family:{FONT};font-size:15px;line-height:1.7;color:#0e2233;-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;">

  <!-- Outer wrapper table for centering -->
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#eaf1f5;">
    <tr>
      <td align="center" style="padding:40px 16px 50px 16px;">

        <!--[if mso]><table role="presentation" width="680" cellpadding="0" cellspacing="0" border="0" align="center"><tr><td><![endif]-->
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="max-width:680px;margin:0 auto;background-color:#ffffff;border:1px solid #d4e1e8;">

          <!-- ========== HEADER ========== -->
          <tr>
            <td style="padding:34px 36px 24px 36px;background-color:#f3f9f8;border-bottom:2px solid #0d8f84;">
              <!-- Eyebrow badge -->
              <table role="presentation" cellpadding="0" cellspacing="0" border="0">
                <tr>
                  <td style="background-color:#e6f4f2;padding:5px 14px;font-size:11px;font-weight:bold;color:#094d47;letter-spacing:0.08em;font-family:{FONT};">
                    &#9679;&nbsp; {course_name}
                  </td>
                </tr>
              </table>
              <!-- Title -->
              <h1 style="margin:16px 0 8px 0;font-size:23px;font-weight:bold;line-height:1.35;color:#091e2f;font-family:{FONT};">{subject}</h1>
              <!-- Recipient info -->
              <p style="margin:0;font-size:14px;line-height:1.7;color:#4a6a82;font-family:{FONT};">{name}&#xB2D8; &middot; {company_dept} &middot; {position}</p>
            </td>
          </tr>

          <!-- ========== BODY ========== -->
          <tr>
            <td style="padding:30px 36px 10px 36px;">
              <p style="margin:0 0 18px 0;font-size:15px;font-weight:bold;color:#094d47;font-family:{FONT};">{name}&#xB2D8; &#xC548;&#xB155;&#xD558;&#xC138;&#xC694;.</p>
              {body_html}
            </td>
          </tr>

          <!-- ========== INFO CARDS (2x2 table) ========== -->
          <tr>
            <td style="padding:0 36px 10px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
                <tr>
                  <!-- Card 1: Core Needs -->
                  <td width="50%" valign="top" style="padding:0 6px 12px 0;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xD575;&#xC2EC; &#xB2C8;&#xC988; (분석)</p>
                          <p style="margin:0;font-size:13px;line-height:1.7;color:#0e2233;font-family:{FONT};">{core_needs}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <!-- Card 2: Desired Outcome -->
                  <td width="50%" valign="top" style="padding:0 0 12px 6px;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xAD50;&#xC721;&#xC744; &#xD64C;&#xD574; &#xC5B7;&#xACE0; &#xC17D;&#xC740; &#xAC83;</p>
                          <p style="margin:0;font-size:13px;line-height:1.7;color:#0e2233;font-family:{FONT};">{desired_outcome}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
                <tr>
                  <!-- Card 3: Psychology State -->
                  <td width="50%" valign="top" style="padding:0 6px 12px 0;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; &#xC124;&#xBB38; &#xAE30;&#xBC18; &#xC1EC;&#xB9AC; &#xC0C1;&#xD3C4;</p>
                          <p style="margin:0;font-size:13px;line-height:1.7;color:#0e2233;font-family:{FONT};">{psychology_state}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <!-- Card 4: Level & Obstacle -->
                  <td width="50%" valign="top" style="padding:0 0 12px 6px;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f7fafa;border:1px solid #dae4ec;">
                      <tr>
                        <td style="padding:16px 18px;">
                          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#0d8f84;letter-spacing:0.06em;font-family:{FONT};">&#9654; AI &#xD65C;&#xC6A9;&#xC131; &amp; &#xC8FC;&#xC9C0; &#xC7A5;&#xC560;&#xBB3C;</p>
                          <p style="margin:0;font-size:13px;line-height:1.7;color:#0e2233;font-family:{FONT};">수준: {ai_level}<br />장애물: {obstacle}</p>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- ========== ENCOURAGEMENT CARD ========== -->
          <tr>
            <td style="padding:4px 36px 16px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff9eb;border:1px solid #fce8b3;border-left:4px solid #f59e0b;">
                <tr>
                  <td style="padding:16px 20px;">
                    <p style="margin:0 0 6px 0;font-size:11px;font-weight:bold;color:#b45309;letter-spacing:0.08em;font-family:{FONT};">&#10024; &#xC6B4;&#xC601;&#xBCF8;&#xBD80;&#xC758; &#xAE0D;&#xB824; &amp; &#xC751;&#xC6D0; &#xBA54;&#xC15C;&#xC9CC;</p>
                    <p style="margin:0;font-size:14px;line-height:1.8;color:#78350f;font-weight:bold;font-family:{FONT};">{encouragement}</p>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- ========== NEXT ACTION ========== -->
          <tr>
            <td style="padding:0 36px 20px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#0d8f84;">
                <tr>
                  <td style="padding:20px 24px;">
                    <p style="margin:0 0 6px 0;font-size:11px;font-weight:bold;letter-spacing:0.1em;color:#b8e6e0;font-family:{FONT};">PRE-SEMINAR CHECKLIST</p>
                    <p style="margin:0;font-size:15px;line-height:1.8;font-weight:bold;color:#ffffff;font-family:{FONT};">{next_action}</p>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- ========== SIGNATURE ========== -->
          <tr>
            <td style="padding:0 36px 36px 36px;">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-top:1px solid #dae4ec;">
                <tr>
                  <td style="padding:20px 0 0 0;font-size:13px;line-height:1.9;color:#4a6a82;font-family:{FONT};">
                    감사합니다.<br />
                    {sender}<br />
                    {sender_org}
                  </td>
                </tr>
              </table>
            </td>
          </tr>

        </table>
        <!--[if mso]></td></tr></table><![endif]-->

      </td>
    </tr>
  </table>

</body>
</html>
"""


def write_index_html(
    index_path: Path,
    rows: list[dict[str, str]],
    generated_rows: list[dict[str, str] | None],
    preview_dir: Path | None,
    args: argparse.Namespace,
) -> None:
    cards: list[str] = []

    for row, generated in zip(rows, generated_rows):
        if generated is None:
            cards.append(
                f"""
        <article class="result-card pending">
          <div class="card-top">
            <div>
              <p class="card-name">{html.escape(row.get("name") or "미상")}</p>
              <p class="card-meta">{html.escape(row.get("company") or "-")} {html.escape(row.get("department") or "-")} | {html.escape(row.get("position") or "-")}</p>
            </div>
            <span class="status pending">Pending</span>
          </div>
        </article>
"""
            )
            continue

        raw_rel = generated.get("feedback_html_path", "")
        href = ""
        if raw_rel and preview_dir:
            file_name = Path(raw_rel).name
            href = f"{preview_dir.name}/{file_name}"

        preview_link = (
            f'<a class="action-btn" href="{html.escape(href)}" target="_blank">이메일 미리보기 HTML ↗</a>'
            if href
            else '<span class="action-btn disabled">미리보기 없음</span>'
        )

        cards.append(
            f"""
        <article class="result-card">
          <div class="card-top">
            <div>
              <p class="card-name">{html.escape(row.get("name") or "미상")} <span class="company-tag">{html.escape(row.get("company") or "-")}</span></p>
              <p class="card-meta">{html.escape(row.get("department") or "-")} &middot; {html.escape(row.get("position") or "-")} | 경력: {html.escape(row.get("experience") or "-")}</p>
            </div>
            <span class="status ok">{html.escape(generated.get("feedback_tone_check") or "ok")}</span>
          </div>

          <div class="card-section">
            <p class="section-label">이메일 제목</p>
            <p class="subject-text">{html.escape(generated.get("feedback_subject") or "-")}</p>
          </div>

          <div class="card-section">
            <p class="section-label">맞춤 안내 메시지</p>
            <p class="message-text">{html.escape(generated.get("feedback_message") or "-")}</p>
          </div>

          <div class="card-section">
            <p class="section-label">✨ 격려 &amp; 응원 메시지</p>
            <p class="encouragement-text">{html.escape(generated.get("feedback_encouragement") or "-")}</p>
          </div>

          <div class="meta-grid">
            <div>
              <p class="section-label">🎯 핵심 니즈 (추론)</p>
              <p class="detail-text">{html.escape(row.get("core_needs") or "-")}</p>
            </div>
            <div>
              <p class="section-label">💡 교육을 통해 얻고 싶은 것</p>
              <p class="detail-text">{html.escape(row.get("desired_outcome") or "-")}</p>
            </div>
            <div>
              <p class="section-label">🧠 심리 상태 (추론)</p>
              <p class="detail-text">{html.escape(row.get("psychology_state") or "-")}</p>
            </div>
            <div>
              <p class="section-label">📋 사전 준비사항 (Pre-checklist)</p>
              <p class="detail-text bold">{html.escape(generated.get("feedback_next_action") or "-")}</p>
            </div>
          </div>

          <div class="card-actions">
            {preview_link}
          </div>
        </article>
"""
        )

    completed_count = sum(item is not None for item in generated_rows)
    rendered_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    document_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{html.escape(args.course_name)} 사전 니즈 대시보드</title>
  <style>
    :root {{
      --bg: #0b1320;
      --card-bg: #132238;
      --border: #1e3556;
      --text: #e6f1fc;
      --muted: #8aa4c4;
      --accent: #00d2c8;
      --accent-glow: rgba(0, 210, 200, 0.15);
      --success: #10b981;
      --pending: #f59e0b;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      background: var(--bg);
      color: var(--text);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      padding: 36px 20px;
      line-height: 1.6;
    }}
    .container {{ max-width: 1100px; margin: 0 auto; }}
    header {{
      background: linear-gradient(135deg, #132238 0%, #0d1829 100%);
      border: 1px solid var(--border);
      border-radius: 16px;
      padding: 28px 32px;
      margin-bottom: 28px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.3);
    }}
    .badge {{
      display: inline-block;
      background: var(--accent-glow);
      color: var(--accent);
      border: 1px solid var(--accent);
      padding: 4px 12px;
      border-radius: 20px;
      font-size: 12px;
      font-weight: 600;
      margin-bottom: 12px;
    }}
    h1 {{ font-size: 26px; font-weight: 700; color: #fff; margin-bottom: 8px; }}
    .subtitle {{ color: var(--muted); font-size: 14px; }}
    .stats-bar {{
      display: flex;
      gap: 20px;
      margin-top: 20px;
      padding-top: 18px;
      border-top: 1px solid var(--border);
    }}
    .stat-item {{ font-size: 13px; color: var(--muted); }}
    .stat-value {{ font-size: 18px; font-weight: 700; color: var(--accent); margin-top: 2px; }}

    .grid {{ display: grid; grid-template-columns: 1fr; gap: 20px; }}
    .result-card {{
      background: var(--card-bg);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 24px;
      transition: transform 0.2s, border-color 0.2s;
    }}
    .result-card:hover {{
      transform: translateY(-2px);
      border-color: var(--accent);
    }}
    .card-top {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 16px; }}
    .card-name {{ font-size: 18px; font-weight: 700; color: #fff; }}
    .company-tag {{ font-size: 13px; font-weight: 500; color: var(--accent); margin-left: 8px; }}
    .card-meta {{ font-size: 13px; color: var(--muted); margin-top: 4px; }}
    .status {{
      font-size: 11px;
      font-weight: 700;
      padding: 4px 10px;
      border-radius: 6px;
      text-transform: uppercase;
    }}
    .status.ok {{ background: rgba(16, 185, 129, 0.15); color: var(--success); border: 1px solid var(--success); }}
    .status.pending {{ background: rgba(245, 158, 11, 0.15); color: var(--pending); border: 1px solid var(--pending); }}

    .card-section {{ margin-bottom: 16px; }}
    .section-label {{ font-size: 11px; font-weight: 700; text-transform: uppercase; color: var(--muted); letter-spacing: 0.05em; margin-bottom: 4px; }}
    .subject-text {{ font-size: 15px; font-weight: 600; color: #fff; }}
    .message-text {{ font-size: 14px; color: #cbd5e1; background: rgba(0,0,0,0.2); padding: 12px 14px; border-radius: 8px; border-left: 3px solid var(--accent); }}
    .encouragement-text {{ font-size: 14px; color: #fde68a; background: rgba(245, 158, 11, 0.1); padding: 12px 14px; border-radius: 8px; border-left: 3px solid #f59e0b; font-weight: 500; }}

    .meta-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
      margin-top: 16px;
      padding-top: 16px;
      border-top: 1px solid var(--border);
    }}
    .detail-text {{ font-size: 13px; color: #94a3b8; }}
    .detail-text.bold {{ color: var(--accent); font-weight: 600; }}

    .card-actions {{ margin-top: 18px; display: flex; justify-content: flex-end; }}
    .action-btn {{
      display: inline-block;
      background: var(--accent);
      color: #0b1320;
      font-size: 13px;
      font-weight: 700;
      padding: 8px 16px;
      border-radius: 8px;
      text-decoration: none;
      transition: opacity 0.2s;
    }}
    .action-btn:hover {{ opacity: 0.9; }}
    .action-btn.disabled {{ background: var(--border); color: var(--muted); cursor: not-allowed; }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <span class="badge">AI-GENERATED NEEDS DASHBOARD</span>
      <h1>{html.escape(args.course_name)} 참가자 사전 니즈 안내 대시보드</h1>
      <p class="subtitle">참가자 핵심니즈, 심리상태, 학습희망사항을 반영한 개인화 HTML 이메일 안내문 모음</p>
      <div class="stats-bar">
        <div class="stat-item">전체 대상자<div class="stat-value">{len(rows)}명</div></div>
        <div class="stat-item">생성 완료<div class="stat-value">{completed_count}명</div></div>
        <div class="stat-item">생성 일시<div class="stat-value">{rendered_at}</div></div>
      </div>
    </header>

    <div class="grid">
      {"".join(cards)}
    </div>
  </div>
</body>
</html>
"""
    index_path.write_text(document_html, encoding="utf-8")


def write_manifest(
    dst_path: Path,
    input_fieldnames: list[str],
    rows: list[dict[str, str]],
    generated_rows: list[dict[str, str] | None],
) -> None:
    output_fieldnames = list(input_fieldnames)
    for field in OUTPUT_FIELDS:
        if field not in output_fieldnames:
            output_fieldnames.append(field)

    with dst_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=output_fieldnames)
        writer.writeheader()

        for row, generated in zip(rows, generated_rows):
            raw_dict = dict(row.get("_raw") or {})
            if generated:
                raw_dict.update(generated)
            else:
                for field in OUTPUT_FIELDS:
                    raw_dict.setdefault(field, "")

            cleaned_dict = {}
            for k, v in raw_dict.items():
                if isinstance(v, str) and k == "feedback_html":
                    cleaned_dict[k] = re.sub(r'[\r\n]+', ' ', v)
                else:
                    cleaned_dict[k] = v
            writer.writerow(cleaned_dict)


def write_output_xlsx(
    xlsx_path: Path,
    input_fieldnames: list[str],
    rows: list[dict[str, str]],
    generated_rows: list[dict[str, str] | None],
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        print("openpyxl 패키지가 없어 XLSX 파일을 생성하지 않습니다.")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "니즈분석_메일안내"

    output_fieldnames = list(input_fieldnames)
    for field in OUTPUT_FIELDS:
        if field not in output_fieldnames:
            output_fieldnames.append(field)

    header_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="0D8F84", end_color="0D8F84", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="DAE4EC"),
    )
    cell_font = Font(name="맑은 고딕", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)
    single_line_align = Alignment(vertical="top", wrap_text=False)
    even_fill = PatternFill(start_color="F4FAFA", end_color="F4FAFA", fill_type="solid")

    for col_idx, field in enumerate(output_fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(output_fieldnames)).column_letter}1"

    for row_idx, (row, generated) in enumerate(zip(rows, generated_rows), start=2):
        raw_dict = dict(row.get("_raw") or {})
        if generated:
            raw_dict.update(generated)
        else:
            for field in OUTPUT_FIELDS:
                raw_dict.setdefault(field, "")

        is_even = (row_idx % 2 == 0)

        for col_idx, field in enumerate(output_fieldnames, start=1):
            val = raw_dict.get(field, "")
            if field == "feedback_html" and isinstance(val, str):
                val = re.sub(r'[\r\n]+', ' ', val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if is_even:
                cell.fill = even_fill

            if field in ("feedback_html", "feedback_html_path", "회사정보출처"):
                cell.alignment = single_line_align
            else:
                cell.alignment = cell_align

    for col_idx, field in enumerate(output_fieldnames, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if field in ("feedback_message", "feedback_html"):
            ws.column_dimensions[col_letter].width = 50
        elif field in ("feedback_subject", "핵심니즈_추론", "심리상태_추론", "교육을통해얻고싶은것_추정", "주요사업"):
            ws.column_dimensions[col_letter].width = 35
        elif field in ("feedback_encouragement", "feedback_next_action", "추론근거", "Q3_학습희망주제"):
            ws.column_dimensions[col_letter].width = 30
        elif field in ("수료증이메일", "회사정보출처", "feedback_html_path"):
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 16

    wb.save(str(xlsx_path))
    return True



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate personalized HRD AI seminar emails from CSV.")
    parser.add_argument(
        "input_csv",
        nargs="?",
        default="sample_data/HR_AI_교육참가자_니즈분석_20260722.csv",
        help="Input CSV file path",
    )
    parser.add_argument("--mode", choices=["glm", "mock"], default="mock", help="Execution mode (mock or glm)")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="GLM model name")
    parser.add_argument("--temperature", type=float, default=0.7, help="Temperature for generation")
    parser.add_argument("--workers", type=int, default=1, help="Parallel worker count")
    parser.add_argument("--limit", type=int, default=None, help="Limit row count for testing")
    parser.add_argument("--course-name", default=DEFAULT_COURSE_NAME, help="Course name")
    parser.add_argument("--sender-name", default=DEFAULT_SENDER_NAME, help="Sender name")
    parser.add_argument("--sender-org", default=DEFAULT_SENDER_ORG, help="Sender organization")
    parser.add_argument("--subject-prefix", default=DEFAULT_SUBJECT_PREFIX, help="Subject prefix")
    parser.add_argument("--output-csv", default=None, help="Custom output CSV path")
    parser.add_argument("--no-html-preview", action="store_true", help="Disable HTML file saving")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output without progress load")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    src_path = Path(args.input_csv).resolve()

    if not src_path.exists():
        print(f"오류: 입력 파일 '{src_path}'을(를) 찾을 수 없습니다.")
        sys.exit(1)

    dst_path = Path(args.output_csv).resolve() if args.output_csv else output_path_for(src_path)
    progress_path = progress_path_for(dst_path)
    preview_dir = None if args.no_html_preview else preview_dir_for(dst_path)
    index_html_path = index_html_path_for(dst_path)
    xlsx_path = xlsx_path_for(dst_path)

    api_key = os.getenv("GLM_API_KEY", "").strip()
    if args.mode == "glm" and not api_key:
        print("오류: --mode glm 사용 시 GLM_API_KEY 환경변수가 필요합니다.")
        sys.exit(1)

    rows, fieldnames, _ = read_input_rows(src_path)
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]

    total_rows = len(rows)
    digest = source_digest(rows, fieldnames)

    print(f"Input rows: {total_rows}")
    print(f"Mode: {args.mode}")
    print(f"Workers: {args.workers}")

    if preview_dir and not preview_dir.exists():
        preview_dir.mkdir(parents=True, exist_ok=True)

    if args.overwrite:
        generated_rows: list[dict[str, str] | None] = [None] * total_rows
    else:
        generated_rows = load_progress(progress_path, src_path, total_rows, digest)

    completed_before = sum(item is not None for item in generated_rows)
    print(f"Generating email guides for {total_rows} rows (completed: {completed_before})...")

    def process_index(idx: int) -> tuple[int, dict[str, str] | None]:
        row = rows[idx]
        name = row.get("name") or f"Row-{idx+1}"
        company = row.get("company") or ""
        print(f"Processing row {idx+1}/{total_rows}: {company} {name}".strip())

        res = generate_feedback(row, args, api_key)
        if res is None:
            return idx, None

        html_content = render_html(row, res, args)
        html_rel_path = ""

        if preview_dir:
            safe_name = sanitize_filename(f"{idx+1:03d}_{company}_{name}")
            html_file = preview_dir / f"{safe_name}.html"
            html_file.write_text(html_content, encoding="utf-8")
            html_rel_path = str(html_file.relative_to(dst_path.parent))

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        output_data = {
            "feedback_subject": res["subject"],
            "feedback_message": res["message"],
            "feedback_encouragement": res["encouragement"],
            "feedback_next_action": res["next_action"],
            "feedback_tone_check": res["tone_check"],
            "feedback_html": html_content,
            "feedback_html_path": html_rel_path,
            "generation_status": "SUCCESS",
            "generation_engine": f"{args.mode}:{args.model}" if args.mode == "glm" else "mock",
            "generated_at": timestamp,
        }
        return idx, output_data

    pending_indices = [i for i, item in enumerate(generated_rows) if item is None]

    if pending_indices:
        if args.workers > 1:
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures = {executor.submit(process_index, i): i for i in pending_indices}
                for future in as_completed(futures):
                    idx, result = future.result()
                    if result:
                        generated_rows[idx] = result
                        save_progress(progress_path, src_path, total_rows, generated_rows, digest)
        else:
            for idx in pending_indices:
                _, result = process_index(idx)
                if result:
                    generated_rows[idx] = result
                    save_progress(progress_path, src_path, total_rows, generated_rows, digest)

    write_manifest(dst_path, fieldnames, rows, generated_rows)
    write_index_html(index_html_path, rows, generated_rows, preview_dir, args)

    try:
        if write_output_xlsx(xlsx_path, fieldnames, rows, generated_rows):
            print(f"Saved XLSX to {xlsx_path}")
    except Exception as exc:
        print(f"XLSX 생성 중 경고: {exc}")

    print(f"Finished! Saved CSV to {dst_path}")
    if preview_dir:
        print(f"HTML previews: {preview_dir}")
    print(f"HTML index dashboard: {index_html_path}")


if __name__ == "__main__":
    main()
